import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook

def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    if archivo:
        entrada_archivo.delete(0, tk.END)
        entrada_archivo.insert(0, archivo)
        cargar_columnas(archivo)

def cargar_columnas(archivo):
    try:
        wb = load_workbook(archivo)
        hoja = wb.active
        encabezados = []
        for col in hoja.iter_cols(1, hoja.max_column, 1, 1):  # Solo fila 1
            encabezado = col[0].value
            if encabezado:
                encabezados.append(encabezado)
        lista_columnas.delete(0, tk.END)
        for encabezado in encabezados:
            lista_columnas.insert(tk.END, encabezado)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar el archivo:\n{e}")

def sumar_columna():
    archivo = entrada_archivo.get()
    seleccion = lista_columnas.curselection()
    if not archivo or not seleccion:
        messagebox.showwarning("Atención", "Selecciona un archivo y una columna.")
        return

    columna_nombre = lista_columnas.get(seleccion)
    try:
        wb = load_workbook(archivo)
        hoja = wb.active

        # Buscar índice de la columna seleccionada
        idx_columna = None
        for i, celda in enumerate(hoja[1], start=1):  # Fila 1
            if celda.value == columna_nombre:
                idx_columna = i
                break

        if idx_columna is None:
            raise ValueError("Columna no encontrada.")

        # Calcular suma manual
        suma = 0
        for fila in hoja.iter_rows(min_row=2, min_col=idx_columna, max_col=idx_columna):
            valor = fila[0].value
            if isinstance(valor, (int, float)):
                suma += valor

        # Agregar fórmula debajo de los datos
        letra_columna = chr(64 + idx_columna)  # A, B, C...
        ultima_fila = hoja.max_row + 1
        formula = f"=SUM({letra_columna}2:{letra_columna}{ultima_fila - 1})"
        hoja.cell(row=ultima_fila, column=idx_columna, value=formula)

        wb.save(archivo)
        messagebox.showinfo("Resultado", f"La suma es: {suma}\nTambién se ha agregado como fórmula en el archivo.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo procesar la columna:\n{e}")

# Interfaz con Tkinter
ventana = tk.Tk()
ventana.title("Suma de columna Excel")

tk.Label(ventana, text="Archivo Excel:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
entrada_archivo = tk.Entry(ventana, width=50)
entrada_archivo.grid(row=0, column=1, padx=5, pady=5)
tk.Button(ventana, text="Seleccionar", command=seleccionar_archivo).grid(row=0, column=2, padx=5, pady=5)

tk.Label(ventana, text="Columnas disponibles:").grid(row=1, column=0, padx=5, pady=5, sticky="ne")
lista_columnas = tk.Listbox(ventana, height=10, width=50)
lista_columnas.grid(row=1, column=1, padx=5, pady=5, columnspan=2)

tk.Button(ventana, text="Sumar columna seleccionada", command=sumar_columna).grid(row=2, column=1, pady=10)

ventana.mainloop()